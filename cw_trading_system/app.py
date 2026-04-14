# app.py
import time
import streamlit as st
import pandas as pd
from io import BytesIO

# Optional imports for Phase 3 features
try:
    import plotly.graph_objects as go
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from cw_trading_system.data.positions import Portfolio, CWPosition, HedgePosition
from cw_trading_system.data.position_store import load_portfolio
from cw_trading_system.data.market_data import MarketDataService
from cw_trading_system.data.market_data_scheduler import MarketDataScheduler

from cw_trading_system.engine.risk_engine import calculate_portfolio_risk
from cw_trading_system.engine.pnl_engine import calculate_pnl
from cw_trading_system.engine.stress_engine import stress_test_grid
from cw_trading_system.engine.hedge_engine import generate_hedge_actions
from cw_trading_system.engine.pricing_engine import price_cw
from cw_trading_system.engine.portfolio_hedge_engine import aggregate_greeks, hedge_by_underlying
from cw_trading_system.engine.issuance_risk_engine import evaluate_issuance_risk
from cw_trading_system.engine.issuance_engine import evaluate_issuance, scan_strikes
from cw_trading_system.engine.strike_optimizer import optimize_strikes, generate_strikes
from cw_trading_system.engine.simulation_engine import simulate_grid
from cw_trading_system.engine.hedging_simulation import simulate_delta_hedge
from cw_trading_system.engine.portfolio_monte_carlo import run_portfolio_mc
from cw_trading_system.engine.monte_carlo_engine import run_monte_carlo
from cw_trading_system.engine.capital_allocation_engine import allocate_capital
from cw_trading_system.engine.trade_execution_engine import trade_execution_engine
from cw_trading_system.engine.reconciliation_engine import reconciliation_engine
from cw_trading_system.engine.market_making_engine import market_making_engine, TradingMode
from cw_trading_system.engine.auto_trading_worker import auto_trading_worker, TradingTask
from cw_trading_system.database import get_session
from cw_trading_system.database.models import Trade

from datetime import date
from cw_trading_system.config.settings import STRESS_SHOCKS, BROKER_CONFIG, VN_MARKET_CONFIG, CW_DATA_CONFIG, BANKING_CONFIG, HEDGE_POLICY
from cw_trading_system.brokers.banking_client import get_capital_overview, check_margin_health, check_capital_adequacy

# 🔐 SECURITY INTEGRATION - Import security utilities
from cw_trading_system.utils.security import input_validator, api_rate_limiter


# =========================
# UTILITY FUNCTIONS
# =========================

def normalize_ticker(ticker: str) -> str:
    """Auto-format ticker to proper format (HPG -> HPG.VN)"""
    if not ticker:
        return ""
    ticker = ticker.upper().strip()
    if not ticker.endswith('.VN'):
        ticker = f"{ticker}.VN"
    return ticker

def get_available_tickers() -> list:
    """Get list of available tickers for dropdown"""
    return ["HPG", "MWG", "VHM", "VIC", "VNM", "GAS", "BID", "CTG", "MBB", "TPB"]

# =========================
# HEALTH CHECK ENDPOINT
# =========================

def health_check() -> dict:
    """
    Health check endpoint for load balancers and monitoring.
    Returns system status and key metrics.
    """
    try:
        # Check database connectivity
        session = get_session()
        session.close()

        # Check market data service
        market_data = MarketDataService()
        test_ticker = "HPG.VN"
        spot = market_data.get_spot_price(test_ticker)

        # Check portfolio loading
        cw_positions, hedge_positions = load_portfolio()
        portfolio = Portfolio(cw_positions, hedge_positions)

        # Check risk calculation
        risk = calculate_portfolio_risk(portfolio)

        return {
            "status": "healthy",
            "timestamp": time.time(),
            "services": {
                "database": "ok",
                "market_data": "ok",
                "portfolio": "ok",
                "risk_engine": "ok"
            },
            "metrics": {
                "portfolio_positions": len(portfolio.cw_positions) + len(portfolio.hedge_positions),
                "total_delta": risk['total']['delta'],
                "total_gamma": risk['total']['gamma']
            }
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": time.time(),
            "error": str(e),
            "services": {
                "database": "error",
                "market_data": "error",
                "portfolio": "error",
                "risk_engine": "error"
            }
        }

# =========================
# PHASE 3: EXPORT FUNCTIONS
# =========================

def export_to_excel(risk, pnl, positions_cw, positions_hedge):
    """Export portfolio data to Excel workbook"""
    if not HAS_OPENPYXL:
        return None
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = "CW Trading System - Portfolio Report"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws_summary[f'A{row}'] = "Portfolio Metrics"
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_summary[f'A{row}'] = "Net Delta"
    ws_summary[f'B{row}'] = risk['total']['delta']
    
    row += 1
    ws_summary[f'A{row}'] = "Gamma"
    ws_summary[f'B{row}'] = risk['total']['gamma']
    
    row += 1
    ws_summary[f'A{row}'] = "Vega"
    ws_summary[f'B{row}'] = risk['total']['vega']
    
    row += 1
    ws_summary[f'A{row}'] = "Theta"
    ws_summary[f'B{row}'] = risk['total']['theta']
    
    row += 1
    ws_summary[f'A{row}'] = "Total P&L"
    ws_summary[f'B{row}'] = pnl['total_pnl']
    ws_summary[f'B{row}'].font = Font(bold=True)
    
    # Sheet 2: Positions
    ws_pos = wb.create_sheet("Positions")
    df_pos = pd.DataFrame([{
        "Type": "CW",
        "Underlying": p.underlying,
        "Ticker": p.ticker,
        "Quantity": p.cw_qty,
        "Strike": p.strike,
        "Expiry": p.expiry,
        "Issue Price": p.issue_price
    } for p in positions_cw])
    
    for r_idx, row in enumerate(df_pos.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws_pos.cell(row=r_idx+1, column=c_idx, value=value)
    
    # Sheet 3: Risk by Underlying
    ws_risk = wb.create_sheet("Risk Analysis")
    df_risk_by_u = pd.DataFrame([{
        "Underlying": u,
        "Delta": r["delta"],
        "Gamma": r["gamma"],
        "Vega": r["vega"],
        "Theta": r["theta"]
    } for u, r in risk["by_underlying"].items()])
    
    for r_idx, row in enumerate(df_risk_by_u.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws_risk.cell(row=r_idx+1, column=c_idx, value=value)
    
    # Convert to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def create_stress_heatmap(stress_results):
    """Create heatmap visualization of stress test results"""
    if not HAS_PLOTLY:
        return None
    
    try:
        # Prepare data
        spot_shocks = sorted(set(k[0] for k in stress_results.keys()))
        vol_shocks = sorted(set(k[1] for k in stress_results.keys()))
        
        z_values = []
        for s_shock in spot_shocks:
            row = []
            for v_shock in vol_shocks:
                pnl = stress_results.get((s_shock, v_shock), {}).get('total_pnl', 0)
                row.append(pnl)
            z_values.append(row)
        
        fig = go.Figure(data=go.Heatmap(
            z=z_values,
            x=[f"{v*100:+.0f}%" for v in vol_shocks],
            y=[f"{s*100:+.0f}%" for s in spot_shocks],
            colorscale='RdYlGn',
            zmid=0,
            text=[[f"${z:,.0f}" for z in row] for row in z_values],
            texttemplate="%{text}",
            textfont={"size": 10},
            colorbar=dict(title="P&L ($)")
        ))
        
        fig.update_layout(
            title="Stress Test P&L Heatmap",
            xaxis_title="Vol Shock",
            yaxis_title="Spot Shock",
            height=500
        )
        
        return fig
    except Exception as e:
        st.warning(f"Error creating heatmap: {e}")
        return None

# =========================
# SIDEBAR NAVIGATION (MOVED LATER)
# =========================

# =========================
# INIT
# =========================

st.set_page_config(layout="wide")
st.title("📊 CW Trading Desk Dashboard (Issuer)")

st.divider()

# --- Portfolio ---
cw_positions, hedge_positions = load_portfolio()
portfolio = Portfolio(cw_positions, hedge_positions)

md = MarketDataService()

if "market_data_scheduler" not in st.session_state:
    st.session_state["market_data_scheduler"] = MarketDataScheduler()

scheduler = st.session_state["market_data_scheduler"]
if not getattr(scheduler, "is_running", False):
    scheduler.start()

# =========================
# CALCULATIONS
# =========================

risk = calculate_portfolio_risk(portfolio, md)
pnl = calculate_pnl(portfolio, md)

spot_shocks = STRESS_SHOCKS
vol_shocks = [-0.10, 0.0, 0.10]

stress = stress_test_grid(portfolio, md, spot_shocks, vol_shocks)
hedge_actions = generate_hedge_actions(risk)

# =========================
# SYSTEM STATUS BANNER (ENHANCED)
# =========================

status_col1, status_col2, status_col3, status_col4, status_col5, status_col6, status_col7, status_col8, status_col9 = st.columns(9)

with status_col1:
    positions_str = f"{len(cw_positions)} CW" if cw_positions else "No CW"
    st.metric("📊 Positions", positions_str)

with status_col2:
    delta_status = "✅ OK" if abs(risk['total']['delta']) < 200_000 else "⚠️ HIGH"
    st.metric("Delta Status", delta_status)

with status_col3:
    pnl_color = "✅ Green" if pnl['total_pnl'] > 0 else "❌ Red"
    st.metric("P&L Status", pnl_color)

with status_col4:
    try:
        spot = md.get_spot("HPG.VN" if cw_positions else "HPG.VN")
        st.metric("Market", f"${spot:.1f}")
    except:
        st.metric("Market", "N/A")

with status_col5:
    broker_status = "✅ Enabled" if BROKER_CONFIG.ocbs_enabled else "⚠️ Disabled"
    st.metric("Broker API", broker_status)

with status_col6:
    # VN Market Data Status
    vn_enabled = any([
        VN_MARKET_CONFIG.vndirect_enabled,
        VN_MARKET_CONFIG.ssi_enabled,
        VN_MARKET_CONFIG.fts_enabled
    ])
    vn_status = "✅ VN APIs" if vn_enabled else "⚠️ Yahoo Only"
    st.metric("VN Market", vn_status)

with status_col7:
    # CW Data Status
    cw_status = "✅ Enabled" if CW_DATA_CONFIG.cw_data_enabled else "⚠️ Disabled"
    st.metric("CW Data", cw_status)

with status_col8:
    # Banking Status
    banking_enabled = any([
        BANKING_CONFIG.banks["vietcombank"]["enabled"],
        BANKING_CONFIG.banks["techcombank"]["enabled"],
        BANKING_CONFIG.banks["vietinbank"]["enabled"]
    ])
    banking_status = "✅ Enabled" if banking_enabled else "⚠️ Disabled"
    st.metric("Banking", banking_status)

with status_col9:
    alerts_count = len(risk.get("breaches", []))
    alert_status = f"{alerts_count} active" if alerts_count > 0 else "None"
    st.metric("Alerts", alert_status)

# =========================
# ACTIVE ALERTS BANNER (NEW)
# =========================

try:
    from cw_trading_system.engine.monitoring_engine import monitoring_engine
    current_alerts = monitoring_engine.get_dashboard_alerts()
except:
    current_alerts = []

if current_alerts or risk.get("breaches", []):
    st.warning("🚨 **ACTIVE RISK ALERTS** - Review immediately")
    with st.expander("📋 View Alert Details", expanded=True):
        for alert in current_alerts:
            st.error(f"  ⚠️ {alert}")
        for breach in risk.get("breaches", []):
            st.error(f"  ⚠️ {breach}")
else:
    st.success("✅ All systems normal - Portfolio within risk limits")

# Portfolio Status
if len(cw_positions) == 0 and len(hedge_positions) == 0:
    st.info("📭 **Portfolio Empty** - Ready to issue new CW positions. Use the 'Issuance & Optimization' tab to start.")

# =========================
# REFRESH CONTROL
# =========================

col1, col2, col3 = st.columns(3)

with col1:
    auto_refresh = st.checkbox("🔄 Auto Refresh", value=False)

with col2:
    refresh_interval = st.selectbox(
        "Interval (sec)",
        [5, 10, 30, 60],
        index=1
    )

with col3:
    if st.button("🔁 Manual Refresh"):
        st.rerun()

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "Overview",
    "Risk & Monitoring",
    "Issuance & Optimization",
    "Simulation",
    "Portfolio & Hedging",
    "Trade Execution",
    "Capital",
    "Banking & Capital"
])

# =========================
# SIDEBAR NAVIGATION (NOW AFTER CALCULATIONS)
# =========================

st.sidebar.title("🚀 Quick Navigation")

# Portfolio Status
st.sidebar.subheader("📊 Portfolio Status")
if len(cw_positions) == 0:
    st.sidebar.info("📭 Empty - Ready to issue")
else:
    st.sidebar.success(f"✅ {len(cw_positions)} CW positions")

# Quick Actions
st.sidebar.subheader("⚡ Quick Actions")
if st.sidebar.button("📸 Record Snapshot", key="sidebar_snapshot"):
    from cw_trading_system.data.monitor_store import record_snapshot
    record_snapshot(risk, pnl)
    st.sidebar.success("Snapshot recorded!")

if st.sidebar.button("🔄 Refresh Data", key="sidebar_refresh"):
    st.rerun()

# Risk Summary
st.sidebar.subheader("⚠️ Risk Summary")
delta_status = "✅ OK" if abs(risk['total']['delta']) < 200_000 else "⚠️ HIGH"
st.sidebar.metric("Delta", delta_status)
pnl_status = "✅ Green" if pnl['total_pnl'] > 0 else "❌ Red"
st.sidebar.metric("P&L", pnl_status)

# Navigation Links
st.sidebar.subheader("📋 Go To")
nav_options = {
    "Overview": "📊",
    "Risk & Monitoring": "⚠️",
    "Issuance & Optimization": "🏗️",
    "Simulation": "🧪",
    "Portfolio & Hedging": "🛡️",
    "Trade Execution": "💱",
    "Capital": "💼"
}

st.sidebar.markdown("**📋 Tabs:**")
st.sidebar.write("Use the tab buttons at the top to navigate between sections.")

# =========================
# SIDEBAR SETTINGS (PHASE 3)
# =========================

st.sidebar.divider()
st.sidebar.subheader("⚙️ Settings")

# Theme selector
theme_option = st.sidebar.selectbox(
    "Theme",
    options=["Light", "Dark", "Auto"],
    help="Change dashboard appearance"
)

# Export section
st.sidebar.subheader("📥 Export")
if st.sidebar.button("📊 Export to Excel"):
    if HAS_OPENPYXL:
        excel_data = export_to_excel(risk, pnl, cw_positions, hedge_positions)
        if excel_data:
            st.sidebar.download_button(
                label="⬇️ Download Report",
                data=excel_data,
                file_name="cw_portfolio_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.sidebar.warning("openpyxl not installed. Install with: pip install openpyxl")

# =========================
# OVERVIEW -TAB1
# =========================
with tab1:
    st.subheader("📌 Portfolio Overview")

    col1, col2, col3, col4, col5 = st.columns(5)

    col1.metric("Net Delta", f"{risk['total']['delta']:,.0f}")
    col2.metric("Gamma", f"{risk['total']['gamma']:,.2f}")
    col3.metric("Vega", f"{risk['total']['vega']:,.0f}")
    col4.metric("Theta", f"{risk['total']['theta']:,.0f}")
    col5.metric("Total P&L", f"{pnl['total_pnl']:,.0f}")

    # Phase 3: Export buttons
    st.divider()
    st.subheader("📥 Export & Reports")
    
    export_col1, export_col2, export_col3 = st.columns(3)
    
    with export_col1:
        if st.button("📊 Export to Excel", key="export_excel_tab1"):
            if HAS_OPENPYXL:
                excel_data = export_to_excel(risk, pnl, cw_positions, hedge_positions)
                if excel_data:
                    st.download_button(
                        label="⬇️ Download Report",
                        data=excel_data,
                        file_name="cw_portfolio_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel"
                    )
            else:
                st.warning("openpyxl library needed: pip install openpyxl")
    
    with export_col2:
        if st.button("📋 Export CSV Data"):
            # Export risk by underlying
            df_risk_export = pd.DataFrame([{
                "Underlying": u,
                "Delta": r["delta"],
                "Gamma": r["gamma"],
                "Vega": r["vega"],
                "Theta": r["theta"]
            } for u, r in risk["by_underlying"].items()])
            
            csv_data = df_risk_export.to_csv(index=False)
            st.download_button(
                label="⬇️ Download CSV",
                data=csv_data,
                file_name="risk_analysis.csv",
                mime="text/csv",
                key="download_csv"
            )
    
    with export_col3:
        st.write("")  # Placeholder

# =========================
# RISK & MONITORING - TAB2
# =========================
with tab2:

    sub1, sub2 = st.tabs(["Risk", "Monitoring"])

    with sub1:
        st.subheader("📊 Risk by Underlying")

        risk_table = []

        for u, r in risk["by_underlying"].items():
            risk_table.append({
                "Underlying": u,
                "Delta": round(r["delta"], 0),
                "Gamma": round(r["gamma"], 2),
                "Vega": round(r["vega"], 0),
                "Theta": round(r["theta"], 0)
            })

        df_risk = pd.DataFrame(risk_table)
        st.dataframe(df_risk, width="stretch")


# =========================
# MONITORING
# =========================
    with sub2:
        st.subheader("📡 Monitoring & Alerts")

        from cw_trading_system.data.monitor_store import record_snapshot, load_logs

        # =========================
        # SNAPSHOT BUTTON
        # =========================

        if st.button("📸 Record Snapshot"):
            record_snapshot(risk, pnl)
            st.success("Snapshot recorded")

        logs = load_logs()

        # =========================
        # RISK TRENDS CHART
        # =========================

        if logs:
            st.subheader("📈 Risk Trends Over Time")

            # Prepare data for charting
            chart_data = pd.DataFrame(logs)
            chart_data['time'] = pd.to_datetime(chart_data['time'])

            # Create line chart for key metrics
            st.line_chart(
                chart_data.set_index('time')[['delta', 'gamma', 'vega', 'theta']],
                width="stretch"
            )

            # P&L chart
            st.subheader("💰 P&L Trends")
            st.line_chart(
                chart_data.set_index('time')[['pnl']],
                width="stretch"
            )

        # =========================
        # TIME SERIES TABLE
        # =========================

        if logs:

            df_logs = pd.DataFrame(logs)
            st.dataframe(df_logs, width="stretch")

            # =========================
            # SIMPLE ALERTS
            # =========================

            st.subheader("🚨 Monitoring Alerts")

            latest = logs[-1]

            # Delta drift alert
            if len(logs) > 1:
                prev = logs[-2]

                delta_change = latest["delta"] - prev["delta"]

                if abs(delta_change) > 100_000:
                    st.warning(f"Large Delta Change: {delta_change:,.0f}")

            # P&L drop alert
            if len(logs) > 1:
                pnl_change = latest["pnl"] - logs[-2]["pnl"]

                if pnl_change < -500_000:
                    st.error(f"P&L Drop: {pnl_change:,.0f}")

        else:
            st.info("No monitoring data yet")

    # =========================
    # ALERT
    # =========================

        st.subheader("🚨 Risk Alerts")

        # Get current alerts from monitoring engine
        from cw_trading_system.engine.monitoring_engine import monitoring_engine
        current_alerts = monitoring_engine.get_dashboard_alerts()

        if current_alerts:
            for alert in current_alerts:
                st.error(f"ACTIVE: {alert}")
        elif risk["breaches"]:
            for b in risk["breaches"]:
                st.warning(b)
        else:
            st.success("All risk within limits")

# =========================
# ISSUANCE
# =========================
with tab3:

    sub1, sub2, sub3 = st.tabs([
        "Issuance",
        "Strike Optimizer",
        "Gamma Check"
    ])

    with sub1:
        st.subheader("🏗️ CW Issuance Strategy")


        # =========================
        # INPUT
        # =========================

        col1, col2, col3 = st.columns(3)

        with col1:
            ticker_raw = st.selectbox(
                "Underlying",
                options=get_available_tickers(),
                index=0,
                help="Select underlying stock"
            )
            ticker = normalize_ticker(ticker_raw)

        with col2:
            expiry = st.date_input("Expiry", value=date(2026, 6, 30))

        with col3:
            strike = st.number_input("Single Strike", value=30.0)

        # =========================
        # SINGLE EVALUATION
        # =========================

        if st.button("Evaluate CW"):
            # 🔐 SECURITY: Validate inputs before processing
            try:
                validated_ticker = input_validator.validate_symbol(ticker_raw)
                validated_strike = input_validator.validate_price(strike, min_price=0.1, max_price=10000.0)
                validated_expiry = expiry.isoformat()  # Date input is already validated by Streamlit
                
                # Rate limiting check
                client_id = "issuance_evaluation"  # Could be user session ID in production
                if not api_rate_limiter.is_allowed(client_id):
                    wait_time = api_rate_limiter.get_wait_time(client_id)
                    st.error(f"⚠️ Rate limit exceeded. Please wait {wait_time:.0f} seconds before trying again.")
                    # Continue execution to show error message
                
                res = evaluate_issuance(md, validated_ticker, validated_strike, validated_expiry)
                pricing = price_cw(md, validated_ticker, validated_strike, validated_expiry)

                st.divider()
                st.subheader("📊 Issuance Decision")

                # Strategy Metrics (improved from JSON)
                st.write("#### Strategy Metrics")
                metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
                with metric_col1:
                    st.metric("Moneyness", f"{res.get('moneyness', 0):.2f}x")
                with metric_col2:
                    st.metric("Vol", f"{res.get('vol', 0):.1%}")
                with metric_col3:
                    st.metric("Risk Score", f"{res.get('risk', 0):.0f}")
                with metric_col4:
                    st.metric("P&L Score", f"{res.get('pnl_score', 0):.2f}")

                # Pricing Details
                st.write("#### Pricing")
                price_col1, price_col2, price_col3, price_col4 = st.columns(4)
                with price_col1:
                    st.metric("Fair Value", f"${pricing.get('fair_value', 0):.2f}")
                with price_col2:
                    st.metric("Issue Price", f"${pricing.get('issuance_price', 0):.2f}")
                with price_col3:
                    edge = pricing.get('issuance_price', 0) - pricing.get('fair_value', 0)
                    st.metric("Edge (Spread)", f"${edge:.2f}")
                with price_col4:
                    edge_pct = (edge / pricing.get('fair_value', 1)) * 100 if pricing.get('fair_value') else 0
                    st.metric("Edge %", f"{edge_pct:.1f}%")

                # Recommendation
                st.divider()
                if edge_pct > 5:
                    st.success("✅ **Excellent opportunity** - Edge > 5%")
                elif edge_pct > 2:
                    st.info("ℹ️ **Good opportunity** - Edge 2-5%")
                else:
                    st.warning("⚠️ **Limited opportunity** - Edge < 2%")
                    
            except ValueError as e:
                st.error(f"❌ **Input Validation Error**: {str(e)}")
                st.info("💡 Please check your inputs and try again.")
            except Exception as e:
                st.error(f"❌ **Processing Error**: {str(e)}")
                st.info("💡 If this persists, please contact support.")


    # =========================
    # AUTO STRIKE GENERATOR
    # =========================
    with sub2:
        st.write("### ⚙️ Auto Strike Generator")

        if st.button("Generate Strikes from Spot"):

            spot = md.get_spot(ticker)

            auto_strikes = [
                round(spot * x, 1)
                for x in [0.9, 0.95, 1.0, 1.05, 1.1, 1.2]
            ]

            strike_range = ",".join([str(k) for k in auto_strikes])

            st.success(f"Generated: {strike_range}")

        # =========================
        # STRIKE SCAN
        # =========================

        st.write("### 🔍 Scan Multiple Strikes")

        col1, col2 = st.columns([3,1])

        with col1:
            strike_range = st.text_input(
                "Strike List (comma separated)",
                value="25,27,30,32,35"
            )

        with col2:
            run_scan = st.button("Run Scan")

        if run_scan:

            try:
                # 🔐 SECURITY: Validate strike range input
                validated_strike_range = input_validator.validate_user_input(strike_range, max_length=200)
                strike_list = [float(x.strip()) for x in validated_strike_range.split(",") if x.strip() != ""]
                
                # Validate each strike price
                validated_strikes = []
                for strike_val in strike_list:
                    validated_strikes.append(input_validator.validate_price(strike_val, min_price=0.1, max_price=10000.0))
                
                # Rate limiting check
                client_id = "strike_scan"
                if not api_rate_limiter.is_allowed(client_id):
                    wait_time = api_rate_limiter.get_wait_time(client_id)
                    st.error(f"⚠️ Rate limit exceeded. Please wait {wait_time:.0f} seconds before trying again.")
                    # Continue execution to show error message

                rows = []

                for K in validated_strikes:

                    res = evaluate_issuance(md, ticker, K, expiry.isoformat())
                    pricing = price_cw(md, ticker, K, expiry.isoformat())

                    rows.append({
                        "Strike": K,
                        "Moneyness": res["moneyness"],
                        "Vol": pricing["sigma"],
                        "Fair Value": pricing["fair_value"],
                        "Issue Price": pricing["issuance_price"],
                        "Edge": pricing["edge"],
                        "Risk": res["risk"]
                    })

                df = pd.DataFrame(rows)

                st.dataframe(df, width="stretch")

            except ValueError as e:
                st.error(f"❌ **Input Validation Error**: {str(e)}")
                st.info("💡 Please check your strike list format (comma-separated numbers) and try again.")
            except Exception as e:
                st.error(f"❌ **Processing Error**: {str(e)}")
                st.info("💡 If this persists, please contact support.")

        # =========================
        # STRIKE OPTIMIZER
        # =========================
        st.write("### 🎯 Strike Optimizer")

        qty = st.number_input("CW Quantity", min_value=1, step=1000, key="opt_qty")
        cr = st.number_input("Conversion Ratio", min_value=0.1, step=0.1, key="opt_cr")
        gamma_limit = st.number_input("Gamma Limit", min_value=1, step=100, key="opt_gamma")
        lambda_risk = st.slider("Risk Aversion (λ)", 0.0, 0.001, 0.0001, key="opt_lambda")

        if st.button("Run Strike Optimization"):
            # 🔐 SECURITY: Validate optimization inputs
            try:
                validated_qty = input_validator.validate_quantity(qty, min_qty=1, max_qty=100_000_000)
                validated_cr = input_validator.validate_percentage(cr, min_val=0.01, max_val=10.0)
                validated_gamma_limit = input_validator.validate_quantity(gamma_limit, min_qty=1, max_qty=1_000_000)
                validated_lambda = input_validator.validate_percentage(lambda_risk, min_val=0.0, max_val=0.01)
                
                # Rate limiting check
                client_id = "strike_optimizer"
                if not api_rate_limiter.is_allowed(client_id):
                    wait_time = api_rate_limiter.get_wait_time(client_id)
                    st.error(f"⚠️ Rate limit exceeded. Please wait {wait_time:.0f} seconds before trying again.")
                    # Continue execution to show error message

                results = optimize_strikes(
                    portfolio,
                    md,
                    ticker,
                    expiry.isoformat(),
                    validated_qty,
                    validated_cr,
                    validated_gamma_limit,
                    validated_lambda
                )

                if results:

                    df = pd.DataFrame(results)
                    st.dataframe(df, width="stretch")

                    best = results[0]

                    st.success(
                        f"Best Strike: {best['strike']} | Score: {round(best['score'],4)}"
                    )

                else:
                    st.error("No feasible strikes under gamma constraint")
                    
            except ValueError as e:
                st.error(f"❌ **Input Validation Error**: {str(e)}")
                st.info("💡 Please check your optimization parameters and try again.")
            except Exception as e:
                st.error(f"❌ **Processing Error**: {str(e)}")
                st.info("💡 If this persists, please contact support.")

    # =========================
    # GAMMA RISK CHECK
    # =========================
    with sub3:
        st.write("### ⚠️ Gamma Risk Check")



        gamma_limit = st.number_input("Gamma Limit", value=10000)

        qty = st.number_input("CW Quantity", value=1_000_000)
        cr = st.number_input("Conversion Ratio", value=1.0, key="gamma_cr")

        if st.button("Check Issuance Risk"):
            # 🔐 SECURITY: Validate gamma risk check inputs
            try:
                validated_gamma_limit = input_validator.validate_quantity(gamma_limit, min_qty=1, max_qty=1_000_000)
                validated_qty = input_validator.validate_quantity(qty, min_qty=1, max_qty=100_000_000)
                validated_cr = input_validator.validate_percentage(cr, min_val=0.01, max_val=10.0)
                
                # Rate limiting check
                client_id = "gamma_risk_check"
                if not api_rate_limiter.is_allowed(client_id):
                    wait_time = api_rate_limiter.get_wait_time(client_id)
                    st.error(f"⚠️ Rate limit exceeded. Please wait {wait_time:.0f} seconds before trying again.")
                    # Continue execution to show error message

                res = evaluate_issuance_risk(
                    portfolio,
                    md,
                    ticker,
                    strike,
                    expiry.isoformat(),
                    validated_qty,
                    validated_cr,
                    validated_gamma_limit
                )

                st.write("Risk Assessment Results:")
                st.json(res)
                    
                col1, col2, col3 = st.columns(3)

                col1.metric("Current Gamma", res["current_gamma"])
                col2.metric("Candidate Gamma", res["candidate_gamma"])
                col3.metric("New Gamma", res["new_gamma"])

                if res["decision"] == "APPROVE":
                    st.success(f"✅ APPROVED: {res['reason']}")
                else:
                    st.error(f"❌ REJECTED: {res['reason']}")
                    
            except ValueError as e:
                st.error(f"❌ **Input Validation Error**: {str(e)}")
                st.info("💡 Please check your risk check parameters and try again.")
            except Exception as e:
                st.error(f"❌ **Processing Error**: {str(e)}")
                st.info("💡 If this persists, please contact support.")



# =========================
# SIMULATION - TAB4
# =========================
with tab4:
    sub1, sub2, sub3, sub4, sub5 = st.tabs([
        "Stress",
        "Grid Simulation",
        "Hedging Simulation",
        "Portfolio Hedge Optimization",
        "Monte Carlo"
    ])

    with sub1:
        st.subheader("📉 Stress Test (Spot × Vol)")

        rows = []

        for s in spot_shocks:
            for v in vol_shocks:
                res = stress[(s, v)]

                rows.append({
                    "Spot Shock": f"{int(s*100)}%",
                    "Vol Shock": f"{int(v*100)}%",
                    "P&L": round(res["total_pnl"], 0)
                })

        df_stress = pd.DataFrame(rows)

        pivot = df_stress.pivot(
            index="Spot Shock",
            columns="Vol Shock",
            values="P&L"
        )

        # Phase 3: Heatmap visualization
        st.subheader("🔥 Heat Map Visualization")
        if HAS_PLOTLY:
            try:
                heatmap_fig = create_stress_heatmap(stress)
                if heatmap_fig:
                    st.plotly_chart(heatmap_fig, use_container_width=True)
            except Exception as e:
                st.warning(f"Could not generate heatmap: {e}")
        else:
            st.info("📊 Heatmap visualization requires plotly. Install with: pip install plotly")

        st.subheader("📊 Detailed Grid")
        st.dataframe(pivot, width="stretch")

# =========================
# GRID SIMULATION
# =========================

    with sub2:
        st.subheader("🧪 Pre-Trade Simulation")

        col1, col2, col3 = st.columns(3)

        with col1:
            ticker_raw = st.selectbox(
                "Underlying",
                options=get_available_tickers(),
                index=0,
                key="sim_ticker_select"
            )
            ticker = normalize_ticker(ticker_raw)

        with col2:
            strike = st.number_input("Strike", value=30.0, key="sim_strike")

        with col3:
            expiry = st.date_input("Expiry", key="sim_expiry")

        # pricing first
        pricing = price_cw(md, ticker, strike, expiry.isoformat())

        issue_price = pricing["issuance_price"]

        st.write(f"**Issuance Price Used:** {issue_price}")

        # shocks
        shocks = [-0.1, -0.05, 0, 0.05, 0.1]

        if st.button("Run Simulation"):

            results = simulate_grid(
                md, ticker, strike, expiry.isoformat(), issue_price, shocks
            )

            df = pd.DataFrame(results)

            st.dataframe(df, width="stretch")


# =========================
# HEDGING SIMULATION
# =========================

    with sub3:

        st.subheader("🔄 Delta Hedging Simulation")

        col1, col2, col3 = st.columns(3)

        with col1:
            ticker_raw = st.selectbox(
                "Underlying",
                options=get_available_tickers(),
                index=0,
                key="hedge_ticker_select"
            )
            ticker = normalize_ticker(ticker_raw)

        with col2:
            strike = st.number_input("Strike", value=30.0, key="hedge_strike")

        with col3:
            expiry = st.date_input("Expiry", key="hedge_expiry")

        shock = st.slider("Final Price Move (%)", -0.2, 0.2, 0.1)

        pricing = price_cw(md, ticker, strike, expiry.isoformat())
        issue_price = pricing["issuance_price"]

        st.write(f"**Issuance Price:** {issue_price}")

        if st.button("Run Hedging Simulation"):

            res = simulate_delta_hedge(
                md,
                ticker,
                strike,
                expiry.isoformat(),
                issue_price,
                shock
            )

            st.write("### 📊 Result")
            st.json(res)

            df_hist = pd.DataFrame(res["history"])
            st.dataframe(df_hist, width="stretch")

# =========================
# PORTFOLIO HEDGE OPTIMIZATION
# =========================

    with sub4:

        st.subheader("📊 Portfolio Hedge Optimization")

        # =========================
        # INPUT PARAMETERS
        # =========================

        col1, col2 = st.columns(2)

        with col1:
            band = st.number_input("Delta Band", value=50000, step=10000)

        with col2:
            st.write(" ")  # spacing
            run_hedge = st.button("Compute Optimal Hedge")

        # =========================
        # RUN HEDGE ENGINE
        # =========================

        if run_hedge:

            # ---- Portfolio level ----
            agg = aggregate_greeks(portfolio, md)

            st.write("### 📌 Portfolio Summary")

            col1, col2 = st.columns(2)
            col1.metric("Total Delta", f"{agg['delta']:,.0f}")
            col2.metric("Total Gamma", f"{agg['gamma']:,.2f}")

            # ---- Underlying breakdown ----
            st.write("### 📊 Hedge Plan by Underlying")

            hedge_plan = hedge_by_underlying(portfolio, md, band)

            rows = []

            for u, res in hedge_plan.items():

                rows.append({
                    "Underlying": u,
                    "Net Delta": round(res["delta"], 0),
                    "Action": res["action"],
                    "Shares to Trade": res["hedge"]
                })

            df = pd.DataFrame(rows)

            st.dataframe(df, width="stretch")

            # =========================
            # INTERPRETATION
            # =========================

            st.write("### 🧠 Interpretation")

            for r in rows:

                if r["Action"] == "TRADE":
                    st.warning(
                        f"{r['Underlying']}: Trade {r['Shares to Trade']} shares to neutralize delta"
                    )
                else:
                    st.success(
                        f"{r['Underlying']}: Within band → no action"
                    )

# =========================
# MONTE CARLO
# =========================
    with sub5:

        st.subheader("🎲 Monte Carlo Risk Simulation")

        col1, col2, col3 = st.columns(3)

        with col1:
            ticker_raw = st.selectbox(
                "Underlying",
                options=get_available_tickers(),
                index=0,
                key="mc_ticker_select"
            )
            ticker = normalize_ticker(ticker_raw)

        with col2:
            strike = st.number_input("Strike", value=30.0, key="mc_strike")

        with col3:
            expiry = st.date_input("Expiry", key="mc_expiry")

        col4, col5 = st.columns(2)

        with col4:
            n_sims = st.number_input("Simulations", value=1000)

        with col5:
            steps = st.number_input("Steps", value=30)

        pricing = price_cw(md, ticker, strike, expiry.isoformat())
        issue_price = pricing["issuance_price"]

        st.write(f"**Issuance Price:** {issue_price}")

        if st.button("Run Monte Carlo"):

            res = run_monte_carlo(
                md,
                ticker,
                strike,
                expiry.isoformat(),
                issue_price,
                int(n_sims),
                int(steps)
            )

            col1, col2, col3, col4 = st.columns(4)

            col1.metric("Mean P&L", round(res["mean"], 2))
            col2.metric("Std Dev", round(res["std"], 2))
            col3.metric("5% Worst", round(res["p5"], 2))
            col4.metric("95% Best", round(res["p95"], 2))

            df = pd.DataFrame(res["all"], columns=["PnL"])
            st.bar_chart(df["PnL"])



        st.subheader("📊 Portfolio Monte Carlo")



        n_sims = st.number_input("Simulations", value=500)

        if st.button("Run Portfolio MC"):

            res = run_portfolio_mc(portfolio, md, int(n_sims))

            col1, col2, col3, col4 = st.columns(4)

            col1.metric("Mean", round(res["mean"], 0))
            col2.metric("Std", round(res["std"], 0))
            col3.metric("5% Worst", round(res["p5"], 0))
            col4.metric("95% Best", round(res["p95"], 0))

            df = pd.DataFrame(res["all"], columns=["PnL"])
            st.histogram(df, x="PnL")

# =========================
# POSITIONS
# =========================
with tab5:
    sub1, sub2, sub3 = st.tabs([
        "Positions",
        "Execution",
        "Hedge Optimization"
    ])

    with sub1:
        st.subheader("📦 Manage CW Positions")

        # =========================
        # ADD NEW CW
        # =========================

        with st.form("add_cw"):

            underlying = st.text_input("Underlying")
            ticker = st.text_input("Ticker (e.g. HPG.VN)")
            qty = st.number_input("Quantity", step=1000)
            strike = st.number_input("Strike")
            expiry = st.date_input("Expiry")
            issue_price = st.number_input("Issue Price")
            cr = st.number_input("Conversion Ratio", value=1.0, key="add_cr")
            sigma = st.number_input("Vol", value=0.30)

            submitted = st.form_submit_button("Add CW")

            if submitted:

                if qty <= 0:
                    st.error("Quantity must be positive")
                else:
                    from cw_trading_system.data.positions import CWPosition
                    from cw_trading_system.data.position_store import add_cw

                    new_pos = CWPosition(
                        underlying,
                        ticker,
                        int(qty),
                        float(cr),
                        float(strike),
                        expiry.isoformat(),
                        float(issue_price),
                        float(sigma)
                    )

                    add_cw(new_pos, cw_positions, hedge_positions)

                    st.success("CW Added")
                    st.rerun()

        # =========================
        # CURRENT POSITIONS
        # =========================

        st.subheader("📋 Current Positions")

        pos_table = []

        for i, p in enumerate(cw_positions):

            # ✅ define T correctly
            T = (date.fromisoformat(p.expiry) - date.today()).days / 365

            vol = md.get_vol(p.ticker, p.strike, T)

            pos_table.append({
                "ID": i,
                "Underlying": p.underlying,
                "Ticker": p.ticker,
                "Qty": p.cw_qty,
                "Strike": p.strike,
                "Expiry": p.expiry,
                "Vol Used": round(vol, 4),  # 👈 NEW
                "CR": p.conversion_ratio
            })

        df_pos = pd.DataFrame(pos_table)
        st.dataframe(df_pos, width="stretch")

        # =========================
        # ACTIONS (SAFE)
        # =========================

        st.subheader("⚙️ Position Actions")

        from cw_trading_system.data.position_store import remove_cw, remove_expired

        if len(cw_positions) > 0:

            remove_idx = st.selectbox(
                "Select CW to remove",
                range(len(cw_positions)),
                format_func=lambda x: f"{x} - {cw_positions[x].ticker}"
            )

            confirm_remove = st.checkbox("Confirm removal")

            if confirm_remove and st.button("❌ Remove Selected CW"):
                remove_cw(remove_idx, cw_positions, hedge_positions)
                st.success("CW Removed")
                st.rerun()

        # Remove expired
        if st.button("🧹 Remove Expired CWs"):
            remove_expired(cw_positions, hedge_positions)
            st.success("Expired CWs removed")
            st.rerun()

# =========================
# TRADING
# =========================
    with sub2:

        st.subheader("🛠️ Hedge Execution")

        from cw_trading_system.data.trade_store import record_trade, load_trades

        # =========================
        # SUGGESTED ACTIONS
        # =========================

        st.write("### Suggested Hedge Actions")

        for action in hedge_actions:

            if action["action"] == "HOLD":
                st.write(f"{action['underlying']}: HOLD")
            else:
                col1, col2 = st.columns([3,1])

                col1.write(
                    f"{action['underlying']}: {action['action']} {action['size']:,}"
                )

                price = col2.number_input(
                    f"Price {action['underlying']}",
                    key=action['underlying']
                )

                if col2.button(f"Execute {action['underlying']}"):
                    record_trade(
                        action["underlying"],
                        action["action"],
                        action["size"],
                        price
                    )
                    st.success("Trade Recorded")
                    st.rerun()

        # =========================
        # TRADE HISTORY
        # =========================

        st.write("### 📜 Trade History")

        trades = load_trades()

        if trades:
            df_trades = pd.DataFrame(trades)
            st.dataframe(df_trades, width="stretch")
        else:
            st.write("No trades yet")

        # =========================
        # P&L
        # =========================
        st.subheader("💰 P&L Breakdown")

        col1, col2 = st.columns(2)

        col1.metric("CW P&L (Short)", f"{pnl['cw_pnl']:,.0f}")
        col2.metric("Hedge P&L", f"{pnl['hedge_pnl']:,.0f}")

# =====================================
# TRADE EXECUTION
# =====================================
with tab6:

    st.subheader("⚡ Trade Execution & Reconciliation")

    # Check if broker integration is enabled
    if not BROKER_CONFIG.ocbs_enabled:
        st.warning("⚠️ OCBS integration is disabled. Configure OCBS settings in .env to enable trade execution.")
    else:
        st.success("✅ OCBS integration enabled")

    # Trade Execution Section
    st.write("### 📈 Execute Trades")

    col1, col2 = st.columns(2)

    with col1:
        st.write("#### CW Issuance")
        cw_ticker = st.text_input("CW Ticker", value="HPG24001", key="cw_ticker")
        cw_underlying = st.text_input("Underlying", value="HPG.VN", key="cw_underlying")
        cw_qty = st.number_input("CW Quantity", value=1000000, key="cw_qty")
        cw_strike = st.number_input("Strike Price", value=30.0, key="cw_strike")
        cw_expiry = st.date_input("Expiry Date", key="cw_expiry")
        cw_conversion_ratio = st.number_input("Conversion Ratio", value=1.0, key="cw_cr")
        cw_issue_price = st.number_input("Issue Price", value=0.5, key="cw_issue_price")

        if st.button("Execute CW Issuance"):
            try:
                issuance_data = {
                    "ticker": cw_ticker,
                    "underlying": cw_underlying,
                    "cw_qty": cw_qty,
                    "strike": cw_strike,
                    "expiry": cw_expiry.isoformat(),
                    "conversion_ratio": cw_conversion_ratio,
                    "issue_price": cw_issue_price
                }

                result = trade_execution_engine.execute_cw_issuance(issuance_data)
                st.success(f"✅ CW Issuance executed! Order ID: {result['order_id']}")
                st.json(result)

            except Exception as e:
                st.error(f"❌ CW Issuance failed: {e}")

    with col2:
        st.write("#### Hedge Trade")
        hedge_underlying = st.text_input("Underlying", value="HPG.VN", key="hedge_underlying")
        hedge_side = st.selectbox("Side", ["buy", "sell"], key="hedge_side")
        hedge_qty = st.number_input("Quantity", value=1000000, key="hedge_qty")
        hedge_reason = st.text_input("Reason", value="Delta hedging", key="hedge_reason")

        if st.button("Execute Hedge Trade"):
            try:
                hedge_data = {
                    "underlying": hedge_underlying,
                    "side": hedge_side,
                    "quantity": hedge_qty,
                    "reason": hedge_reason
                }

                result = trade_execution_engine.execute_hedge_trade(hedge_data)
                st.success(f"✅ Hedge trade executed! Order ID: {result['order_id']}")
                st.json(result)

            except Exception as e:
                st.error(f"❌ Hedge trade failed: {e}")

    st.divider()
    st.write("### 🏪 Live Market Making Workflow")
    st.caption("Use Streamlit as the control panel while the pricing and routing logic stays in the backend engine.")

    mode_options = {
        "Manual": TradingMode.MANUAL,
        "Semi-auto": TradingMode.SEMI_AUTO,
        "Auto": TradingMode.AUTO,
    }

    mm_col1, mm_col2 = st.columns([2, 1])

    with mm_col1:
        mm_ticker = st.text_input("CW for Market Making", value=cw_ticker, key="mm_ticker")
        mm_underlying = st.text_input("Underlying for MM", value=cw_underlying, key="mm_underlying")
        mm_mode_label = st.selectbox("MM Mode", list(mode_options.keys()), key="mm_mode_label")
        mm_quote_qty = st.number_input("Quote Size", min_value=1, value=100000, step=10000, key="mm_quote_qty")
        mm_inventory = st.number_input("Current Inventory", value=0, step=10000, key="mm_inventory")
        mm_iv_premium = st.slider("IV Premium (%)", min_value=0.0, max_value=25.0, value=10.0, step=0.5, key="mm_iv_premium")
        mm_spread = st.slider("Bid/Ask Spread (%)", min_value=0.5, max_value=10.0, value=4.0, step=0.5, key="mm_spread")

    with mm_col2:
        mm_state = market_making_engine.get_workflow_state(mm_ticker, mm_underlying)
        st.metric("MM Status", mm_state["market_making_status"].title())
        st.metric("Hedge Status", mm_state["hedge_status"].title())
        st.metric("Kill Switch", "ON" if mm_state["kill_switch"] else "OFF")
        st.caption(mm_state["last_message"])

    preview_col, submit_col, start_col, pause_col, kill_col = st.columns(5)

    if preview_col.button("Preview Quote", key="preview_mm_quote"):
        try:
            live_pricing = price_cw(md, mm_underlying, cw_strike, cw_expiry.isoformat())
            quote = market_making_engine.build_two_sided_quote(
                ticker=mm_ticker,
                underlying=mm_underlying,
                fair_value=live_pricing["fair_value"],
                iv_premium=mm_iv_premium / 100,
                spread_pct=mm_spread / 100,
                quantity=int(mm_quote_qty),
                inventory=int(mm_inventory),
                mode=mode_options[mm_mode_label],
            )
            st.session_state["latest_mm_quote"] = quote
            st.session_state["latest_mm_pricing"] = live_pricing
            st.success("✅ Two-sided quote generated.")
        except Exception as e:
            st.error(f"❌ Quote preview failed: {e}")

    if submit_col.button("Submit Quote", key="submit_mm_quote"):
        try:
            quote = st.session_state.get("latest_mm_quote")
            if not quote:
                live_pricing = price_cw(md, mm_underlying, cw_strike, cw_expiry.isoformat())
                quote = market_making_engine.build_two_sided_quote(
                    ticker=mm_ticker,
                    underlying=mm_underlying,
                    fair_value=live_pricing["fair_value"],
                    iv_premium=mm_iv_premium / 100,
                    spread_pct=mm_spread / 100,
                    quantity=int(mm_quote_qty),
                    inventory=int(mm_inventory),
                    mode=mode_options[mm_mode_label],
                )
            mm_result = market_making_engine.submit_two_sided_quote(quote, send_orders=True)
            st.session_state["latest_mm_result"] = mm_result
            if mm_result["success"]:
                status_msg = "submitted to OCBS" if mm_result.get("broker_enabled") else "simulated because OCBS is disabled in this environment"
                st.success(f"✅ Buy and sell quotes {status_msg}.")
            else:
                st.error("❌ Quote submission blocked by checks.")
        except Exception as e:
            st.error(f"❌ Quote submission failed: {e}")

    if start_col.button("Start Auto MM", key="start_auto_mm"):
        state = market_making_engine.set_market_making_mode(
            ticker=mm_ticker,
            mode=mode_options[mm_mode_label],
            enabled=True,
            underlying=mm_underlying,
        )
        st.success(f"✅ Market making {state['status']} in {state['mode']} mode.")

    if pause_col.button("Pause MM", key="pause_auto_mm"):
        state = market_making_engine.set_market_making_mode(
            ticker=mm_ticker,
            mode=mode_options[mm_mode_label],
            enabled=False,
            underlying=mm_underlying,
        )
        st.info(f"⏸️ Market making {state['status']}.")

    if kill_col.button("Toggle Kill", key="toggle_mm_kill_switch"):
        state = market_making_engine.set_kill_switch(
            ticker=mm_ticker,
            active=not mm_state["kill_switch"],
            underlying=mm_underlying,
        )
        if state["kill_switch"]:
            st.warning("🛑 Kill switch activated. All live trading is blocked.")
        else:
            st.success("✅ Kill switch released.")

    if "latest_mm_quote" in st.session_state:
        st.write("#### Current MM Quote")
        st.dataframe(pd.DataFrame([st.session_state["latest_mm_quote"]]), width="stretch")

    if "latest_mm_result" in st.session_state:
        st.write("#### Last Quote Routing Result")
        st.json(st.session_state["latest_mm_result"])

    workflow_rows = [
        {
            "Step": "1. Price CW",
            "Action": "Calculate fair value + IV premium",
            "Current State": mm_state["market_making_status"],
        },
        {
            "Step": "2. Send Quotes",
            "Action": "Route buy and sell limits to OCBS",
            "Current State": mm_state["market_making_mode"],
        },
        {
            "Step": "3. Hedge Exposure",
            "Action": "Check delta band and generate hedge",
            "Current State": mm_state["hedge_status"],
        },
    ]
    st.dataframe(pd.DataFrame(workflow_rows), width="stretch")

    st.write("### 🛡️ Live Hedging Workflow")

    hedge_mode_options = {
        "Manual": TradingMode.MANUAL,
        "Auto": TradingMode.AUTO,
    }

    hedge_flow_col1, hedge_flow_col2 = st.columns([2, 1])

    with hedge_flow_col1:
        hedge_workflow_mode = st.selectbox("Hedge Mode", list(hedge_mode_options.keys()), key="hedge_workflow_mode")
        live_net_delta = st.number_input("Net Delta to Hedge", value=float(risk["total"]["delta"]), step=10000.0, key="live_net_delta")
        hedge_band = st.number_input("Delta Band", min_value=1.0, value=float(HEDGE_POLICY.delta_band), step=10000.0, key="hedge_band_live")
        hedge_ratio_live = st.slider("Hedge Ratio", min_value=0.1, max_value=1.5, value=float(HEDGE_POLICY.hedge_ratio), step=0.1, key="hedge_ratio_live")
        hedge_min_size = st.number_input("Minimum Hedge Size", min_value=1, value=int(HEDGE_POLICY.min_trade_size), step=1000, key="hedge_min_size_live")
        hedge_workflow_underlying = st.text_input("Hedge Underlying", value=hedge_underlying, key="hedge_workflow_underlying")

    with hedge_flow_col2:
        st.metric("Portfolio Delta", f"{risk['total']['delta']:,.0f}")
        st.metric("Configured Band", f"{hedge_band:,.0f}")
        st.metric("Broker", "OCBS Live" if BROKER_CONFIG.ocbs_enabled else "Simulation")

    suggest_col, execute_col, enable_col, disable_col = st.columns(4)

    if suggest_col.button("Suggest Hedge", key="suggest_hedge_workflow"):
        decision = market_making_engine.evaluate_hedge_need(
            underlying=hedge_workflow_underlying,
            net_delta=live_net_delta,
            delta_band=hedge_band,
            hedge_ratio=hedge_ratio_live,
            min_trade_size=int(hedge_min_size),
        )
        st.session_state["latest_hedge_decision"] = decision
        if decision["requires_hedge"]:
            st.warning(f"⚠️ Hedge required: {decision['side']} {decision['quantity']:,} shares.")
        else:
            st.success("✅ No hedge required under the current rules.")

    if execute_col.button("Execute Hedge", key="execute_hedge_workflow"):
        decision = st.session_state.get("latest_hedge_decision") or market_making_engine.evaluate_hedge_need(
            underlying=hedge_workflow_underlying,
            net_delta=live_net_delta,
            delta_band=hedge_band,
            hedge_ratio=hedge_ratio_live,
            min_trade_size=int(hedge_min_size),
        )
        hedge_result = market_making_engine.execute_hedge_decision(
            ticker=mm_ticker,
            decision=decision,
            auto_execute=hedge_mode_options[hedge_workflow_mode] == TradingMode.AUTO,
        )
        st.session_state["latest_hedge_result"] = hedge_result
        if hedge_result["success"]:
            st.success(hedge_result.get("message", "✅ Hedge workflow processed."))
        else:
            st.error(hedge_result.get("message", "❌ Hedge workflow blocked."))

    if enable_col.button("Enable Auto Hedge", key="enable_auto_hedge"):
        state = market_making_engine.set_hedging_mode(
            ticker=mm_ticker,
            mode=hedge_mode_options[hedge_workflow_mode],
            enabled=True,
            underlying=hedge_workflow_underlying,
        )
        st.success(f"✅ Hedge engine {state['status']} in {state['mode']} mode.")

    if disable_col.button("Pause Hedge", key="pause_auto_hedge"):
        state = market_making_engine.set_hedging_mode(
            ticker=mm_ticker,
            mode=hedge_mode_options[hedge_workflow_mode],
            enabled=False,
            underlying=hedge_workflow_underlying,
        )
        st.info(f"⏸️ Hedge engine {state['status']}.")

    if "latest_hedge_decision" in st.session_state:
        st.write("#### Latest Hedge Decision")
        st.json(st.session_state["latest_hedge_decision"])

    if "latest_hedge_result" in st.session_state:
        st.write("#### Last Hedge Result")
        st.json(st.session_state["latest_hedge_result"])

    st.divider()
    st.write("### 🤖 OCBS Auto-Execution Worker")
    st.caption("Runs periodic quote and hedge cycles in the background. Until the OCBS API is connected, it operates in simulation-ready mode.")

    worker_summary = auto_trading_worker.get_status()
    worker_detail = auto_trading_worker.get_status(mm_ticker, mm_underlying)

    worker_col1, worker_col2 = st.columns([2, 1])

    with worker_col1:
        worker_interval = st.number_input("Worker Interval (sec)", min_value=5, value=30, step=5, key="worker_interval_seconds")
        worker_mm_enabled = st.checkbox("Enable Auto Market Making", value=True, key="worker_mm_enabled")
        worker_hedge_enabled = st.checkbox("Enable Auto Hedging", value=True, key="worker_hedge_enabled")
        use_live_worker_delta = st.checkbox("Use live portfolio delta", value=True, key="worker_use_live_delta")
        worker_delta_override = st.number_input("Delta Override", value=float(risk["total"]["delta"]), step=10000.0, key="worker_delta_override", disabled=use_live_worker_delta)
        worker_notes = st.text_input("Strategy Notes", value="CW live market making strategy", key="worker_notes")

    with worker_col2:
        st.metric("Worker Running", "Yes" if worker_summary["worker_running"] else "No")
        st.metric("Strategies", worker_summary["strategy_count"])
        st.metric("Broker Mode", worker_summary["mode"])

    register_col, cycle_col, start_worker_col, stop_worker_col, remove_col = st.columns(5)

    if register_col.button("Register Strategy", key="register_auto_strategy"):
        try:
            task = TradingTask(
                ticker=mm_ticker,
                underlying=mm_underlying,
                strike=float(cw_strike),
                expiry=cw_expiry.isoformat(),
                market_making_enabled=worker_mm_enabled,
                auto_hedging_enabled=worker_hedge_enabled,
                interval_seconds=int(worker_interval),
                iv_premium=mm_iv_premium / 100,
                spread_pct=mm_spread / 100,
                quote_quantity=int(mm_quote_qty),
                inventory=int(mm_inventory),
                trading_mode=mode_options[mm_mode_label].value,
                delta_band=float(hedge_band),
                hedge_ratio=float(hedge_ratio_live),
                min_trade_size=int(hedge_min_size),
                delta_override=None if use_live_worker_delta else float(worker_delta_override),
                notes=worker_notes,
            )
            registration = auto_trading_worker.register_strategy(task)
            st.session_state["auto_worker_registration"] = registration
            st.success("✅ Auto-execution strategy registered.")
        except Exception as e:
            st.error(f"❌ Failed to register strategy: {e}")

    if cycle_col.button("Run Cycle Now", key="run_auto_cycle_now"):
        try:
            if not worker_detail.get("task"):
                task = TradingTask(
                    ticker=mm_ticker,
                    underlying=mm_underlying,
                    strike=float(cw_strike),
                    expiry=cw_expiry.isoformat(),
                    market_making_enabled=worker_mm_enabled,
                    auto_hedging_enabled=worker_hedge_enabled,
                    interval_seconds=int(worker_interval),
                    iv_premium=mm_iv_premium / 100,
                    spread_pct=mm_spread / 100,
                    quote_quantity=int(mm_quote_qty),
                    inventory=int(mm_inventory),
                    trading_mode=mode_options[mm_mode_label].value,
                    delta_band=float(hedge_band),
                    hedge_ratio=float(hedge_ratio_live),
                    min_trade_size=int(hedge_min_size),
                    delta_override=None if use_live_worker_delta else float(worker_delta_override),
                    notes=worker_notes,
                )
                auto_trading_worker.register_strategy(task)

            cycle_result = auto_trading_worker.run_cycle(mm_ticker, mm_underlying)
            st.session_state["auto_worker_cycle_result"] = cycle_result
            if cycle_result["success"]:
                st.success("✅ Auto-execution cycle completed.")
            else:
                st.error(f"❌ Auto-execution cycle failed: {cycle_result.get('error', 'Unknown error')}")
        except Exception as e:
            st.error(f"❌ Failed to run worker cycle: {e}")

    if start_worker_col.button("Start Worker", key="start_auto_worker"):
        result = auto_trading_worker.start()
        st.success(f"✅ Worker status: {result['status']}")

    if stop_worker_col.button("Stop Worker", key="stop_auto_worker"):
        result = auto_trading_worker.stop()
        st.info(f"⏹️ Worker status: {result['status']}")

    if remove_col.button("Remove Strategy", key="remove_auto_strategy"):
        result = auto_trading_worker.unregister_strategy(mm_ticker, mm_underlying)
        if result["removed"]:
            st.success("✅ Strategy removed from worker.")
        else:
            st.warning("ℹ️ No registered strategy was found to remove.")

    refreshed_worker_detail = auto_trading_worker.get_status(mm_ticker, mm_underlying)
    st.write("#### Worker Status")
    st.json(refreshed_worker_detail)

    if "auto_worker_cycle_result" in st.session_state:
        st.write("#### Last Worker Cycle")
        st.json(st.session_state["auto_worker_cycle_result"])

    # Reconciliation Section
    st.write("### 🔍 Position Reconciliation")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("Reconcile Positions"):
            try:
                result = reconciliation_engine.reconcile_positions()
                discrepancies = result.get("discrepancies", [])

                if discrepancies:
                    st.warning(f"⚠️ Found {len(discrepancies)} discrepancies")
                    df = pd.DataFrame(discrepancies)
                    st.dataframe(df, width="stretch")
                else:
                    st.success("✅ All positions reconciled successfully")

            except Exception as e:
                st.error(f"❌ Reconciliation failed: {e}")

    with col2:
        reconcile_days = st.number_input("Days back for trade reconciliation", value=1, min_value=1)
        if st.button("Reconcile Trades"):
            try:
                result = reconciliation_engine.reconcile_trades(reconcile_days)
                matched = len(result.get("matched", []))
                unmatched_int = len(result.get("unmatched_internal", []))
                unmatched_brk = len(result.get("unmatched_broker", []))

                st.info(f"Matched: {matched}, Unmatched Internal: {unmatched_int}, Unmatched Broker: {unmatched_brk}")

                if unmatched_int > 0 or unmatched_brk > 0:
                    st.warning("⚠️ Trade reconciliation discrepancies found")
                    if unmatched_int > 0:
                        st.write("**Unmatched Internal Trades:**")
                        st.json(result["unmatched_internal"][:5])  # Show first 5
                    if unmatched_brk > 0:
                        st.write("**Unmatched Broker Trades:**")
                        st.json(result["unmatched_broker"][:5])  # Show first 5
                else:
                    st.success("✅ All trades reconciled successfully")

            except Exception as e:
                st.error(f"❌ Trade reconciliation failed: {e}")

    # Recent Trades Section
    st.write("### 📋 Recent Trades")

    try:
        session = get_session()
        recent_trades = session.query(Trade).order_by(Trade.execution_time.desc()).limit(10).all()
        session.close()

        if recent_trades:
            trades_data = [{
                "ID": t.id,
                "Action": t.action.value,
                "Underlying": t.underlying,
                "Quantity": t.quantity,
                "Price": t.price,
                "Time": t.execution_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Status": t.status
            } for t in recent_trades]

            df = pd.DataFrame(trades_data)
            st.dataframe(df, width="stretch")
        else:
            st.info("No recent trades found")

    except Exception as e:
        st.error(f"Failed to load recent trades: {e}")


# =====================================
# CAPITAL ALLOCATION & ROE OPTIMIZATION
# =====================================
with tab7:

    st.subheader("💼 Capital Allocation & ROE Optimization")



    total_capital = st.number_input("Total Capital", value=1_000_000_000)

    ticker_raw = st.selectbox(
        "Underlying",
        options=get_available_tickers(),
        index=0,
        key="capital_ticker_select"
    )
    ticker = normalize_ticker(ticker_raw)
    expiry = st.date_input("Expiry")

    qty = st.number_input("CW Size", value=1_000_000)
    cr = st.number_input("Conversion Ratio", value=1.0, key="capital_cr")

    spot = md.get_spot(ticker)
    strikes = generate_strikes(spot)

    candidates = []

    for K in strikes:
        candidates.append({
            "ticker": ticker,
            "strike": K,
            "expiry": expiry.isoformat(),
            "qty": qty,
            "cr": cr
        })

    if st.button("Run Allocation"):

        res = allocate_capital(
            portfolio,
            md,
            candidates,
            total_capital
        )

        df = pd.DataFrame(res["selected"])
        st.dataframe(df, width="stretch")

        st.metric("Capital Used", round(res["used_capital"], 0))


# =====================================
# BANKING & CAPITAL MANAGEMENT
# =====================================
with tab8:

    st.subheader("🏦 Banking & Capital Management")

    # Banking Status Overview
    st.subheader("📊 Banking Integration Status")

    banking_col1, banking_col2, banking_col3 = st.columns(3)

    with banking_col1:
        vcb_enabled = BANKING_CONFIG.banks["vietcombank"]["enabled"]
        st.metric("Vietcombank", "✅ Enabled" if vcb_enabled else "⚠️ Disabled")

    with banking_col2:
        tcb_enabled = BANKING_CONFIG.banks["techcombank"]["enabled"]
        st.metric("Techcombank", "✅ Enabled" if tcb_enabled else "⚠️ Disabled")

    with banking_col3:
        vtb_enabled = BANKING_CONFIG.banks["vietinbank"]["enabled"]
        st.metric("Vietinbank", "✅ Enabled" if vtb_enabled else "⚠️ Disabled")

    # Capital Overview
    st.subheader("💰 Capital Overview")

    if st.button("🔄 Refresh Capital Data", key="refresh_capital"):
        try:
            capital_overview = get_capital_overview()

            cap_col1, cap_col2, cap_col3, cap_col4 = st.columns(4)

            with cap_col1:
                st.metric("Available Balance",
                         f"${capital_overview['available_balance']:,.0f}")

            with cap_col2:
                margin_health = capital_overview['margin_health']
                st.metric("Margin Available",
                         f"${margin_health.get('total_margin_available', 0):,.0f}")

            with cap_col3:
                pending_amount = capital_overview.get('total_pending_amount', 0)
                st.metric("Pending Settlements",
                         f"${pending_amount:,.0f}")

            with cap_col4:
                utilization = margin_health.get('overall_utilization', 0)
                status = margin_health.get('overall_status', 'UNKNOWN')
                st.metric("Margin Status", f"{status} ({utilization:.1%})")

            # Margin Health Details
            if margin_health.get('margin_call_alerts'):
                st.error("🚨 **MARGIN CALL ALERTS**")
                for alert in margin_health['margin_call_alerts']:
                    st.error(f"  ⚠️ {alert['bank']} - {alert['severity']}")

            # Pending Settlements
            pending_settlements = capital_overview.get('pending_settlements', [])
            if pending_settlements:
                st.warning(f"⚠️ {len(pending_settlements)} pending settlements")

                with st.expander("View Pending Settlements"):
                    df_pending = pd.DataFrame(pending_settlements)
                    st.dataframe(df_pending, width="stretch")

        except Exception as e:
            st.error(f"Failed to load capital data: {e}")
            st.info("💡 Configure banking API credentials in .env file to enable capital monitoring")

    # Capital Adequacy Check
    st.subheader("🔍 Capital Adequacy Check")

    required_amount = st.number_input(
        "Required Amount for Transaction",
        value=1000000,
        min_value=0,
        step=100000
    )

    if st.button("Check Adequacy"):
        try:
            adequacy = check_capital_adequacy(required_amount)

            adequacy_col1, adequacy_col2, adequacy_col3 = st.columns(3)

            with adequacy_col1:
                sufficient = "✅ Sufficient" if adequacy['sufficient'] else "❌ Insufficient"
                st.metric("Capital Check", sufficient)

            with adequacy_col2:
                st.metric("Effective Available",
                         f"${adequacy['effective_available']:,.0f}")

            with adequacy_col3:
                shortfall = adequacy.get('shortfall', 0)
                if shortfall > 0:
                    st.metric("Shortfall", f"${shortfall:,.0f}")
                else:
                    st.metric("Surplus", f"${adequacy['effective_available'] - required_amount:,.0f}")

        except Exception as e:
            st.error(f"Failed to check capital adequacy: {e}")

    # Configuration Info
    st.subheader("⚙️ Configuration")
    st.info("""
    **Banking APIs require configuration in .env file:**

    ```
    # Vietcombank
    VCB_ENABLED=true
    VCB_BASE_URL=https://api.vietcombank.com.vn
    VCB_API_KEY=your-api-key
    VCB_API_SECRET=your-api-secret
    VCB_ACCOUNT_NUMBER=your-account

    # Techcombank
    TCB_ENABLED=true
    TCB_BASE_URL=https://api.techcombank.com.vn
    TCB_API_KEY=your-api-key
    TCB_API_SECRET=your-api-secret
    TCB_ACCOUNT_NUMBER=your-account

    # Vietinbank
    VTB_ENABLED=true
    VTB_BASE_URL=https://api.vietinbank.vn
    VTB_API_KEY=your-api-key
    VTB_API_SECRET=your-api-secret
    VTB_ACCOUNT_NUMBER=your-account
    ```
    """)
